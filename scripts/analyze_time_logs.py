#!/usr/bin/env python3
"""
Time Log Analysis — Operation Darrentan
Parses 7 years (2019-2025) of color-coded Excel time logs.

Each half-hour block with a colored fill = 0.5 hours of work in that category.
Categories are mapped from legends embedded in each file's first week block.
"""

import json
import sys
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

# --- File definitions ---

FILES = {
    2019: {
        'path': '/home/darney/Downloads/time-logs/Time Log 2019.xlsm.xlsx',
        'sheet': 'Sheet1',
        'legend_cols': (13, 14),  # M, N
        'week_rows': 52,  # 1 header + 48 timeslots + 3 footer
        'header_rows': 1,  # just "Time" + dates row
        'date_row_offset': 0,  # dates are in the header row itself
    },
    2020: {'path': '/home/darney/Downloads/time-logs/Time Log 2020.xlsm.xlsx', 'sheet': 'Time Log'},
    2021: {'path': '/home/darney/Downloads/time-logs/Time Log 2021.xlsm.xlsx', 'sheet': 'Time Log'},
    2022: {'path': '/home/darney/Downloads/time-logs/Time Log 2022.xlsm.xlsx', 'sheet': 'Time Log'},
    2023: {'path': '/home/darney/Downloads/time-logs/Time Log 2023.xlsm.xlsx', 'sheet': 'Time Log'},
    2024: {'path': '/home/darney/Downloads/time-logs/Time Log 2024.xlsm.xlsx', 'sheet': 'Time Log'},
    2025: {'path': '/home/darney/Downloads/Time Log 2025.xlsx', 'sheet': 'Time Log'},
}

# Default structure for 2020-2025
for y in range(2020, 2026):
    FILES[y].setdefault('legend_cols', (9, 10))
    FILES[y].setdefault('week_rows', 51)     # "Week" + dates + 48 timeslots + 1 blank
    FILES[y].setdefault('header_rows', 2)     # "Week" row + dates row
    FILES[y].setdefault('date_row_offset', 1) # dates are 1 row below "Week"

# Colors to always exclude (background/formatting, not work)
EXCLUDE_COLORS = {
    'FFBFBFBF',  # Gray in 2019 — off-hours background
    '00000000',  # No fill
}

# PTO color (marks date cells, not time slots)
PTO_COLOR = 'FF92D050'
PTO_COLORS = {'FF92D050', 'FFCCCCCC'}  # 2025 uses FFCCCCCC for PTO

# --- Category classification ---

def classify_category(name, color_rgb):
    """Classify a category into high-level buckets for analysis."""
    name_lower = name.lower().strip()

    # Personal/Life
    if any(k in name_lower for k in ['personal', 'dadmin', 'family time', 'bbc book']):
        return 'personal'

    # PTO
    if 'pto' in name_lower or 'holiday' in name_lower:
        return 'pto'

    # Overhead (Biosero corporate, management, admin, presales, general apps)
    if any(k in name_lower for k in [
        'biosero (admin)', 'biosero admin', 'management', 'email/admin',
        'presales', 'pre-sales', 'general apps', 'post-sales',
        'apps / post-sales', 'general post-sales',
        'portfolio work', 'planner work',
    ]):
        return 'overhead'

    # "Biosero" alone in early years = catch-all work category
    if name_lower == 'biosero':
        return 'overhead'

    # Client project numbers or named accounts = billable
    if any(c.isdigit() for c in name[:5]):  # starts with project number
        return 'billable'
    if any(k in name_lower for k in ['lilly', 'cchmc', 'illumina', 'xenon', 'bms']):
        return 'billable'

    # Fallback — if it has a color and isn't excluded, likely billable
    if color_rgb not in EXCLUDE_COLORS and color_rgb not in PTO_COLORS:
        return 'billable'

    return 'unknown'


def get_cell_color(cell):
    """Extract the fill color RGB from a cell, or None."""
    if not cell.fill or not cell.fill.fgColor:
        return None
    rgb = cell.fill.fgColor.rgb
    if rgb is None:
        return None
    rgb_str = str(rgb)
    if rgb_str == '00000000':
        return None
    return rgb_str


def extract_legend(ws, color_col, label_col, max_rows=55):
    """Extract color→category mapping from the first legend block only.

    Stops at "Total Hours" marker. A row with a color swatch but no label
    is NOT treated as blank (it's an unlabeled legend entry).
    Only keeps the FIRST mapping for each color.
    """
    legend = {}
    found_entries = False
    truly_blank_streak = 0

    for r in range(1, max_rows):
        color_cell = ws.cell(row=r, column=color_col)
        label_cell = ws.cell(row=r, column=label_col)

        # Check for "Total Hours" marker — stop here
        color_val = str(color_cell.value).strip() if color_cell.value else ''
        label_val = str(label_cell.value).strip() if label_cell.value else ''
        if 'total hours' in color_val.lower() or 'total hours' in label_val.lower():
            break
        if 'total' == color_val.lower() or 'total' == label_val.lower():
            break

        rgb = get_cell_color(color_cell)
        label = label_val if label_val and label_val != 'None' and label_val != '#NAME?' else ''

        if rgb and label:
            if rgb not in legend:
                legend[rgb] = label
            found_entries = True
            truly_blank_streak = 0
        elif rgb:
            # Color swatch exists but no label — still part of legend block
            truly_blank_streak = 0
        elif found_entries and not rgb:
            # Truly blank row (no color, no useful label)
            truly_blank_streak += 1
            if truly_blank_streak >= 4:
                break

    return legend


def extract_legend_from_config(wb):
    """Extract legend from 2025-style Config sheet (category + hex code mapping)."""
    if 'Config' not in wb.sheetnames:
        return {}

    ws = wb['Config']
    legend = {}

    for r in range(3, 50):  # skip headers
        cat_cell = ws.cell(row=r, column=1)
        hex_cell = ws.cell(row=r, column=2)

        cat_name = str(cat_cell.value).strip() if cat_cell.value else ''
        hex_val = str(hex_cell.value).strip() if hex_cell.value else ''

        if not cat_name or not hex_val or hex_val == 'None':
            continue

        # Convert #rrggbb to FFrrggbb (openpyxl format)
        hex_val = hex_val.lstrip('#')
        if len(hex_val) == 6:
            rgb = 'FF' + hex_val.upper()
            legend[rgb] = cat_name

    return legend


def parse_year(year, config):
    """Parse a single year's time log and return structured data."""
    wb = openpyxl.load_workbook(config['path'], data_only=True)
    ws = wb[config['sheet']]

    color_col, label_col = config['legend_cols']
    week_rows = config['week_rows']
    header_rows = config['header_rows']
    date_row_offset = config['date_row_offset']

    # Extract legend from Time Log sheet
    legend = extract_legend(ws, color_col, label_col)

    # Supplement with Config sheet if available (2025)
    config_legend = extract_legend_from_config(wb)
    for rgb, name in config_legend.items():
        if rgb not in legend:
            legend[rgb] = name

    print(f"\n{'='*60}")
    print(f"  {year}: {len(legend)} categories in legend")
    for rgb, name in sorted(legend.items(), key=lambda x: x[1]):
        cls = classify_category(name, rgb)
        print(f"    {rgb} = {name} [{cls}]")

    # Parse week blocks
    weeks = []
    row = 1
    max_row = ws.max_row

    while row <= max_row:
        # Find next week header
        cell_a = ws.cell(row=row, column=1)
        cell_val = str(cell_a.value).strip().lower() if cell_a.value else ''

        is_week_header = False
        if year == 2019:
            is_week_header = cell_val == 'time'
        elif year == 2020:
            # 2020 uses "Week 1", "Week 2" etc — combined header with dates
            is_week_header = cell_val.startswith('week ')
        else:
            is_week_header = cell_val == 'week'

        if not is_week_header:
            row += 1
            continue

        # Found a week header — extract dates from the date row
        if year == 2019:
            date_row = row  # dates are on the same row as "Time"
        elif year == 2020:
            date_row = row  # "Week N" row has dates in B-H
        else:
            date_row = row + 1  # dates are on row after "Week"

        week_dates = []
        pto_days = set()
        for col in range(2, 9):  # B-H = 7 days
            date_cell = ws.cell(row=date_row, column=col)
            date_val = date_cell.value
            if isinstance(date_val, datetime):
                week_dates.append(date_val)
            elif isinstance(date_val, str):
                try:
                    week_dates.append(datetime.strptime(date_val, '%Y-%m-%d'))
                except ValueError:
                    week_dates.append(None)
            else:
                week_dates.append(None)

            # Check if this day is PTO (date cell has PTO color)
            date_color = get_cell_color(date_cell)
            if date_color in PTO_COLORS:
                pto_days.add(col - 2)  # 0-indexed day

        # Get week number and date range
        valid_dates = [d for d in week_dates if d is not None]
        if not valid_dates:
            row += week_rows
            continue

        week_start = min(valid_dates)
        week_end = max(valid_dates)

        # Count half-hour blocks by color
        timeslot_start = date_row + 1  # first timeslot row
        timeslot_end = timeslot_start + 48  # 48 half-hour slots

        category_blocks = defaultdict(int)
        total_blocks = 0

        for slot_row in range(timeslot_start, min(timeslot_end, max_row + 1)):
            for col in range(2, 9):  # B-H
                cell = ws.cell(row=slot_row, column=col)
                rgb = get_cell_color(cell)
                if rgb and rgb not in EXCLUDE_COLORS:
                    category_blocks[rgb] += 1
                    total_blocks += 1

        # Convert blocks to hours (each block = 0.5 hours)
        category_hours = {}
        for rgb, blocks in category_blocks.items():
            cat_name = legend.get(rgb, f'Unknown ({rgb})')
            cat_class = classify_category(cat_name, rgb)
            key = (cat_name, cat_class)
            category_hours[key] = category_hours.get(key, 0) + blocks * 0.5

        total_hours = total_blocks * 0.5
        pto_day_count = len(pto_days)

        # Separate work hours (billable + overhead) from personal/pto
        work_hours = sum(hrs for (name, cls), hrs in category_hours.items() if cls in ('billable', 'overhead'))
        billable_hours = sum(hrs for (name, cls), hrs in category_hours.items() if cls == 'billable')
        overhead_hours = sum(hrs for (name, cls), hrs in category_hours.items() if cls == 'overhead')

        weeks.append({
            'week_start': week_start.strftime('%Y-%m-%d'),
            'week_end': week_end.strftime('%Y-%m-%d'),
            'total_hours': total_hours,
            'work_hours': work_hours,
            'billable_hours': billable_hours,
            'overhead_hours': overhead_hours,
            'pto_days': pto_day_count,
            'category_hours': {f"{name}|{cls}": hrs for (name, cls), hrs in category_hours.items()},
        })

        row += week_rows

    wb.close()

    # Compute year summary
    if not weeks:
        return None

    total_weeks = len(weeks)
    # Filter out weeks with zero WORK hours (personal-only or empty)
    active_weeks = [w for w in weeks if w['work_hours'] > 0]
    active_week_count = len(active_weeks)

    all_work_hours = sum(w['work_hours'] for w in weeks)
    all_logged_hours = sum(w['total_hours'] for w in weeks)
    avg_work_hours = all_work_hours / active_week_count if active_week_count else 0

    # Aggregate category hours across the year
    year_categories = defaultdict(float)
    year_classes = defaultdict(float)
    for w in weeks:
        for key, hrs in w['category_hours'].items():
            name, cls = key.rsplit('|', 1)
            year_categories[name] += hrs
            year_classes[cls] += hrs

    # Weekly work stats (billable + overhead only)
    weekly_work = [w['work_hours'] for w in active_weeks]
    max_work_week = max(weekly_work) if weekly_work else 0
    min_work_week = min(weekly_work) if weekly_work else 0

    # PTO days total
    total_pto_days = sum(w['pto_days'] for w in weeks)

    return {
        'year': year,
        'total_weeks': total_weeks,
        'active_weeks': active_week_count,
        'total_hours': round(all_logged_hours, 1),
        'work_hours': round(all_work_hours, 1),
        'avg_hours_per_week': round(avg_work_hours, 1),
        'max_week_hours': round(max_work_week, 1),
        'min_week_hours': round(min_work_week, 1),
        'pto_days': total_pto_days,
        'by_class': {k: round(v, 1) for k, v in sorted(year_classes.items())},
        'by_category': {k: round(v, 1) for k, v in sorted(year_categories.items(), key=lambda x: -x[1])},
        'weeks': weeks,
    }


def compute_negative_space(summary):
    """Calculate the 'life' side — hours NOT at work.

    Uses WORK hours (billable + overhead), not total logged hours.
    This is critical because early years (2019-2020) also logged personal time.
    """
    total_possible = 168  # hours per week
    work_hours = summary['avg_hours_per_week']  # already billable+overhead only
    free_hours = total_possible - work_hours

    # Sleep assumption: 7h/night = 49h/week
    sleep = 49
    waking_free = free_hours - sleep

    # Overhead hours that would disappear as independent
    overhead = summary['by_class'].get('overhead', 0)
    overhead_per_week = overhead / summary['active_weeks'] if summary['active_weeks'] else 0

    # Billable efficiency — what % of work time is client-facing
    billable = summary['by_class'].get('billable', 0)
    work_total = summary['work_hours']
    billable_pct = (billable / work_total * 100) if work_total else 0

    return {
        'avg_work_hours': round(work_hours, 1),
        'avg_free_hours': round(free_hours, 1),
        'waking_free_hours': round(waking_free, 1),
        'overhead_hours_per_week': round(overhead_per_week, 1),
        'overhead_reclaimed_monthly': round(overhead_per_week * 4.33, 1),
        'billable_efficiency': round(billable_pct, 1),
    }


def count_unique_clients(results):
    """Count unique pharma client accounts across all years (exclude overhead/personal)."""
    clients = set()
    overhead_keywords = ['biosero', 'management', 'email', 'admin', 'presales', 'pre-sales',
                         'general apps', 'post-sales', 'portfolio', 'planner', 'personal',
                         'dadmin', 'family', 'pto', 'holiday', 'bbc book', 'unknown']
    for r in results:
        for cat in r['by_category']:
            cat_lower = cat.lower()
            if not any(k in cat_lower for k in overhead_keywords):
                clients.add(cat)
    return clients


def print_analysis(results):
    """Print the comprehensive multi-year analysis."""
    print("\n" + "=" * 80)
    print("  OPERATION DARRENTAN — 7-YEAR TIME LOG ANALYSIS")
    print("  Sell availability, not hours. Charge for what you're worth.")
    print("=" * 80)

    all_active = sum(r['active_weeks'] for r in results)
    all_work = sum(r['work_hours'] for r in results)
    career_avg = all_work / all_active if all_active else 0

    # --- WORK SIDE ---
    print("\n" + "-" * 80)
    print("  THE WORK SIDE: What Was Given")
    print("  (Work hours = billable + overhead; personal time excluded)")
    print("-" * 80)

    print(f"\n{'Year':<6} {'Weeks':>6} {'Work Hrs':>10} {'Avg/Wk':>8} {'Max Wk':>8} {'Min Wk':>8} {'PTO Days':>9}")
    print("-" * 62)
    for r in results:
        print(f"{r['year']:<6} {r['active_weeks']:>6} {r['work_hours']:>10.1f} {r['avg_hours_per_week']:>8.1f} "
              f"{r['max_week_hours']:>8.1f} {r['min_week_hours']:>8.1f} {r['pto_days']:>9}")
    print("-" * 62)
    print(f"{'TOTAL':<6} {all_active:>6} {all_work:>10.1f} {career_avg:>8.1f}")

    # Billable vs overhead trend (percentages based on work hours, not total)
    print(f"\n{'Year':<6} {'Billable':>10} {'Overhead':>10} {'Work Tot':>10} {'Bill %':>8} {'Over %':>8}")
    print("-" * 62)
    for r in results:
        billable = r['by_class'].get('billable', 0)
        overhead = r['by_class'].get('overhead', 0)
        work = r['work_hours']
        bill_pct = (billable / work * 100) if work else 0
        over_pct = (overhead / work * 100) if work else 0
        print(f"{r['year']:<6} {billable:>10.1f} {overhead:>10.1f} {work:>10.1f} {bill_pct:>7.1f}% {over_pct:>7.1f}%")

    # Top work categories (exclude personal/pto)
    print("\n  TOP CLIENT ACCOUNTS (All Years Combined)")
    print("-" * 60)
    all_cats = defaultdict(float)
    for r in results:
        for cat, hrs in r['by_category'].items():
            all_cats[cat] += hrs
    work_cats = {k: v for k, v in all_cats.items()
                 if not any(x in k.lower() for x in ['personal', 'dadmin', 'family', 'pto', 'holiday'])}
    for cat, hrs in sorted(work_cats.items(), key=lambda x: -x[1])[:20]:
        print(f"  {hrs:>8.1f}h  {cat}")

    # Peak capacity evidence (based on work hours, not total)
    print("\n  PEAK CAPACITY EVIDENCE")
    print("-" * 60)
    for r in results:
        big_weeks = [w for w in r['weeks'] if w['work_hours'] >= 50]
        if big_weeks:
            print(f"  {r['year']}: {len(big_weeks)} weeks >= 50h work (max {r['max_week_hours']:.1f}h)")
        else:
            big40 = [w for w in r['weeks'] if w['work_hours'] >= 40]
            print(f"  {r['year']}: {len(big40)} weeks >= 40h work (max {r['max_week_hours']:.1f}h)")

    # --- LIFE SIDE ---
    print("\n" + "-" * 80)
    print("  THE LIFE SIDE: The Negative Space")
    print("  168 hours/week - work hours = everything else")
    print("-" * 80)

    print(f"\n{'Year':<6} {'Work/Wk':>8} {'Free/Wk':>8} {'Wake Free':>10} {'Over/Wk':>8} {'Bill Eff':>9}")
    print("-" * 62)
    for r in results:
        ns = compute_negative_space(r)
        print(f"{r['year']:<6} {ns['avg_work_hours']:>8.1f} {ns['avg_free_hours']:>8.1f} "
              f"{ns['waking_free_hours']:>10.1f} {ns['overhead_hours_per_week']:>8.1f} "
              f"{ns['billable_efficiency']:>8.1f}%")

    # Oliver timeline
    print("\n  LIFE TIMELINE OVERLAY")
    print("-" * 60)
    events = [
        (2019, "Jul-Dec only — time tracking begins, Biosero catch-all"),
        (2020, "Pandemic wedding, WFH shift, world upside down"),
        (2021, "Post-pandemic stabilization, fertility journey begins"),
        (2022, "Peak intensity — 10+ concurrent pharma accounts"),
        (2023, "Oliver born September — everything changes"),
        (2024, "Full year as father, workload vs family tension peaks"),
        (2025, "'Lateral move' → recalibration → Operation Darrentan"),
    ]
    for y, event in events:
        r = next((x for x in results if x['year'] == y), None)
        if r:
            ns = compute_negative_space(r)
            print(f"\n  {y}: {r['avg_hours_per_week']:.1f}h/wk work → "
                  f"{ns['waking_free_hours']:.1f}h/wk for life")
            print(f"       {event}")

    # Pre/post Oliver comparison (using work hours)
    pre_oliver = [r for r in results if r['year'] < 2023]
    post_oliver = [r for r in results if r['year'] >= 2024]

    if pre_oliver and post_oliver:
        pre_avg = sum(r['work_hours'] for r in pre_oliver) / sum(r['active_weeks'] for r in pre_oliver)
        post_avg = sum(r['work_hours'] for r in post_oliver) / sum(r['active_weeks'] for r in post_oliver)
        pre_overhead = sum(r['by_class'].get('overhead', 0) for r in pre_oliver) / sum(r['active_weeks'] for r in pre_oliver)
        post_overhead = sum(r['by_class'].get('overhead', 0) for r in post_oliver) / sum(r['active_weeks'] for r in post_oliver)

        print(f"\n  PRE-OLIVER (2019-2022):  {pre_avg:.1f}h/wk work, {pre_overhead:.1f}h/wk overhead")
        print(f"  POST-OLIVER (2024-2025): {post_avg:.1f}h/wk work, {post_overhead:.1f}h/wk overhead")
        diff = pre_avg - post_avg
        if diff > 0:
            print(f"  SHIFT: -{diff:.1f}h/wk — fatherhood demanded space, work compressed")
            print(f"  Overhead drop: -{pre_overhead - post_overhead:.1f}h/wk — already shedding company tax")

    # --- PRICING STRATEGY ---
    print("\n" + "-" * 80)
    print("  PRICING STRATEGY: Sell Availability, Not Hours")
    print("-" * 80)

    career_billable = sum(r['by_class'].get('billable', 0) for r in results)
    career_overhead = sum(r['by_class'].get('overhead', 0) for r in results)
    career_work = all_work
    career_billable_pct = (career_billable / career_work * 100) if career_work else 0
    overhead_per_week = career_overhead / all_active if all_active else 0
    billable_per_week = career_billable / all_active if all_active else 0

    unique_clients = count_unique_clients(results)

    print(f"\n  WHAT DISAPPEARS AS INDEPENDENT:")
    print(f"    Career overhead: {career_overhead:.0f}h ({overhead_per_week:.1f}h/wk)")
    print(f"    Annual overhead: ~{overhead_per_week * 52:.0f}h/year → reclaimed as family time")
    print(f"    That's {overhead_per_week * 52 / 8:.0f} full 8-hour days/year back")
    print(f"    Categories: Biosero admin, management, presales, general apps/post-sales")

    print(f"\n  RETAINER MATH (LIFE-FIRST):")
    print(f"    Monthly nut: ~$5,200 (bills $4k + insurance $800 + therapy/meds $400)")
    print(f"    At $10k/mo retainer from ONE client:")
    print(f"      → Covers nut with 48% margin for taxes + savings")
    print(f"    At market value ($200-250/hr):")
    print(f"      → Your {billable_per_week:.0f}h/wk billable avg = ${billable_per_week * 200:,.0f}-${billable_per_week * 250:,.0f}/wk delivered")
    print(f"      → Charging $10k/mo (~$2.5k/wk) = client pays ~12% of market value")

    print(f"\n  MINIMUM VIABLE SCHEDULE:")
    print(f"    Floor: 20h/wk billable × $125/hr = $10k/mo (3 focused days)")
    print(f"    Target: 25-30h/wk across 2 retainers = $20k/mo ($240k/yr)")
    print(f"    Career avg: {career_avg:.1f}h/wk TOTAL work — minus overhead = {billable_per_week:.0f}h/wk billable")
    print(f"    As independent: ALL {career_avg:.1f}h becomes billable (overhead gone)")
    print(f"    Even 2025 'disengaged': {results[-1]['avg_hours_per_week']:.1f}h/wk still above target")

    print(f"\n  WHAT THE TIME LOGS PROVE TO CLIENTS:")
    print(f"    - {all_active} weeks of documented output over 7 years")
    print(f"    - {career_billable_pct:.0f}% billable efficiency (of work hours)")
    print(f"    - {len(unique_clients)} distinct pharma/biotech client accounts")
    print(f"    - Max sustained: {max(r['max_week_hours'] for r in results):.0f}h work weeks")
    print(f"    - Overhead Biosero charged clients for? You did it unpaid")

    return results


def save_results(results, output_path):
    """Save analysis results as JSON for API integration."""
    output = {
        'generated': datetime.now().isoformat(),
        'years': [],
        'career_summary': {},
    }

    all_active = sum(r['active_weeks'] for r in results)
    all_work = sum(r['work_hours'] for r in results)
    career_avg = all_work / all_active if all_active else 0

    career_billable = sum(r['by_class'].get('billable', 0) for r in results)
    career_overhead = sum(r['by_class'].get('overhead', 0) for r in results)
    career_personal = sum(r['by_class'].get('personal', 0) for r in results)

    unique_clients = count_unique_clients(results)

    output['career_summary'] = {
        'total_weeks': all_active,
        'total_work_hours': round(all_work, 1),
        'avg_work_hours_per_week': round(career_avg, 1),
        'total_billable': round(career_billable, 1),
        'total_overhead': round(career_overhead, 1),
        'total_personal': round(career_personal, 1),
        'billable_efficiency': round(career_billable / all_work * 100, 1) if all_work else 0,
        'overhead_pct': round(career_overhead / all_work * 100, 1) if all_work else 0,
        'overhead_per_week': round(career_overhead / all_active, 1) if all_active else 0,
        'billable_per_week': round(career_billable / all_active, 1) if all_active else 0,
        'unique_client_accounts': len(unique_clients),
    }

    for r in results:
        year_data = {
            'year': r['year'],
            'active_weeks': r['active_weeks'],
            'work_hours': r['work_hours'],
            'avg_hours_per_week': r['avg_hours_per_week'],
            'max_week_hours': r['max_week_hours'],
            'min_week_hours': r['min_week_hours'],
            'pto_days': r['pto_days'],
            'by_class': r['by_class'],
            'by_category': r['by_category'],
            'negative_space': compute_negative_space(r),
        }
        output['years'].append(year_data)

    with open(output_path, 'w') as f:
        json.dump(output, f, indent=2)

    print(f"\n  Results saved to {output_path}")


if __name__ == '__main__':
    results = []
    for year in sorted(FILES.keys()):
        config = FILES[year]
        print(f"\nProcessing {year}...")
        summary = parse_year(year, config)
        if summary:
            results.append(summary)
        else:
            print(f"  WARNING: No data for {year}")

    analysis = print_analysis(results)

    # Save JSON output
    output_path = Path('/home/darney/projects/proxmox-homelab/scripts/time_log_analysis.json')
    save_results(results, output_path)
